"""
FastAPI Sales Analytics Backend
Handles CRUD operations for sales data and provides analytics endpoints
"""
from fastapi import FastAPI, Depends, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

from database import engine, SessionLocal, Base
from models import SalesRecord
from schemas import SalesRecordCreate, SalesRecordUpdate, SalesRecordResponse

# Create database tables
Base.metadata.create_all(bind=engine)

# Initialize FastAPI app
app = FastAPI(
    title="Sales Analytics API",
    description="API for sales data management and analytics",
    version="1.0.0"
)

# CORS configuration for React frontend
allowed_origins = [
    "http://localhost:3000",
    "http://localhost:5173",
    "http://localhost:8000",
]

# Add production frontend URL if available
frontend_url = os.getenv("FRONTEND_URL", "").strip()
if frontend_url:
    allowed_origins.append(frontend_url)
else:
    # Fallback for development
    allowed_origins.extend([
        "https://*.vercel.app",
        "https://*.render.com",
    ])

app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Dependency to get database session
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ==================== CRUD ENDPOINTS ====================

@app.post("/api/sales", response_model=SalesRecordResponse, status_code=201)
def create_sales_record(record: SalesRecordCreate, db: Session = Depends(get_db)):
    """
    Create a new sales record.
    
    Request body:
    - region: str (East, West, North, South)
    - sales_amount: float
    - price: float
    - profit: float
    - discount: float
    - date: str (YYYY-MM-DD)
    - month: str (January - December)
    """
    db_record = SalesRecord(**record.dict())
    db.add(db_record)
    db.commit()
    db.refresh(db_record)
    return db_record


@app.get("/api/sales", response_model=list[SalesRecordResponse])
def get_all_sales(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    region: str = Query(None),
    month: str = Query(None),
    db: Session = Depends(get_db)
):
    """
    Get all sales records with optional filtering and pagination.
    
    Query parameters:
    - skip: number of records to skip (pagination)
    - limit: maximum number of records to return
    - region: filter by region (optional)
    - month: filter by month (optional)
    """
    query = db.query(SalesRecord)
    
    if region:
        query = query.filter(SalesRecord.region == region)
    if month:
        query = query.filter(SalesRecord.month == month)
    
    records = query.offset(skip).limit(limit).all()
    return records


@app.get("/api/sales/{record_id}", response_model=SalesRecordResponse)
def get_sales_record(record_id: int, db: Session = Depends(get_db)):
    """Get a specific sales record by ID."""
    record = db.query(SalesRecord).filter(SalesRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="Sales record not found")
    return record


@app.put("/api/sales/{record_id}", response_model=SalesRecordResponse)
def update_sales_record(
    record_id: int,
    record_update: SalesRecordUpdate,
    db: Session = Depends(get_db)
):
    """Update an existing sales record."""
    db_record = db.query(SalesRecord).filter(SalesRecord.id == record_id).first()
    if not db_record:
        raise HTTPException(status_code=404, detail="Sales record not found")
    
    update_data = record_update.dict(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_record, key, value)
    
    db.commit()
    db.refresh(db_record)
    return db_record


@app.delete("/api/sales/{record_id}", status_code=204)
def delete_sales_record(record_id: int, db: Session = Depends(get_db)):
    """Delete a sales record."""
    db_record = db.query(SalesRecord).filter(SalesRecord.id == record_id).first()
    if not db_record:
        raise HTTPException(status_code=404, detail="Sales record not found")
    
    db.delete(db_record)
    db.commit()
    return None


# ==================== ANALYTICS ENDPOINTS ====================

@app.get("/api/analytics/summary")
def get_analytics_summary(db: Session = Depends(get_db)):
    """Get key metrics: total sales, total profit, average discount."""
    records = db.query(SalesRecord).all()
    
    if not records:
        return {
            "total_sales": 0,
            "total_profit": 0,
            "average_discount": 0,
            "record_count": 0
        }
    
    total_sales = sum(r.sales_amount for r in records)
    total_profit = sum(r.profit for r in records)
    average_discount = sum(r.discount for r in records) / len(records)
    
    return {
        "total_sales": round(total_sales, 2),
        "total_profit": round(total_profit, 2),
        "average_discount": round(average_discount, 2),
        "record_count": len(records)
    }


@app.get("/api/analytics/by-region")
def get_sales_by_region(db: Session = Depends(get_db)):
    """Get sales grouped by region."""
    regions = ["East", "West", "North", "South"]
    result = []
    
    for region in regions:
        records = db.query(SalesRecord).filter(SalesRecord.region == region).all()
        total_sales = sum(r.sales_amount for r in records)
        result.append({
            "region": region,
            "sales": round(total_sales, 2),
            "count": len(records)
        })
    
    return result


@app.get("/api/analytics/by-month")
def get_sales_by_month(db: Session = Depends(get_db)):
    """Get sales grouped by month."""
    months = ["January", "February", "March", "April", "May", "June",
              "July", "August", "September", "October", "November", "December"]
    result = []
    
    for month in months:
        records = db.query(SalesRecord).filter(SalesRecord.month == month).all()
        total_sales = sum(r.sales_amount for r in records)
        if total_sales > 0:  # Only include months with sales
            result.append({
                "month": month,
                "sales": round(total_sales, 2),
                "count": len(records)
            })
    
    return result


@app.get("/api/analytics/profit-by-region")
def get_profit_by_region(db: Session = Depends(get_db)):
    """Get profit distributed by region."""
    regions = ["East", "West", "North", "South"]
    result = []
    
    for region in regions:
        records = db.query(SalesRecord).filter(SalesRecord.region == region).all()
        total_profit = sum(r.profit for r in records)
        result.append({
            "region": region,
            "profit": round(total_profit, 2),
            "count": len(records)
        })
    
    return result


# ==================== EXPORT ENDPOINTS ====================

@app.get("/api/export/excel")
def export_to_excel(db: Session = Depends(get_db)):
    """Export all sales records to an Excel file."""
    records = db.query(SalesRecord).all()
    
    if not records:
        raise HTTPException(status_code=404, detail="No records to export")
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Data"
    
    # Define headers
    headers = ["ID", "Region", "Sales Amount", "Price", "Profit", "Discount", "Date", "Month", "Created At"]
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data rows
    for record in records:
        ws.append([
            record.id,
            record.region,
            record.sales_amount,
            record.price,
            record.profit,
            record.discount,
            record.date.strftime("%Y-%m-%d"),
            record.month,
            record.created_at.strftime("%Y-%m-%d %H:%M:%S")
        ])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 20
    
    # Save file
    filename = f"sales_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join("exports", filename)
    
    # Create exports directory if it doesn't exist
    os.makedirs("exports", exist_ok=True)
    
    wb.save(filepath)
    return {"filename": filename, "path": filepath, "count": len(records)}


# ==================== HEALTH CHECK ====================

@app.get("/health")
def health_check():
    """Health check endpoint."""
    return {"status": "healthy", "timestamp": datetime.now().isoformat()}


if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", 8000))
    host = os.getenv("HOST", "0.0.0.0")
    uvicorn.run(app, host=host, port=port)
